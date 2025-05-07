import pandas as pd
from openpyxl import load_workbook
import sys

if len(sys.argv) > 1:
        
    sheet_name = sys.argv[1] # = 'Hs15_logsum'

    print("Sheet name: ", sheet_name)

    excel_file_path_petr = "C:/Users/HajduPe/OneDrive - Helsingin Seudun liikenne - Kuntayhtymä/Estimointi/EstimointikoneHelmet5 – kopio - wPR.xlsx"
    excel_file_path = excel_file_path_petr #'EstimointikoneHelmet5_esimerkki.xlsx'

    # Load the existing Excel file
    workbook = load_workbook(excel_file_path)

    vastaavuudet = workbook["Vastaavuudet"]
    sheet_lookup = ""
    for row in vastaavuudet.iter_rows(min_row=1, max_col=2):
        if row[0].value == sheet_name:
            sheet_lookup = row[1].value
            sheet_lookup = sheet_lookup.split("\\")[-1][:-4]
    
    # File paths
    alogit_run_petr_path = "C:/Users/HajduPe/H4_estimointi/Helmet4/helmet_estimation/ohjaus/"
    text_file_path = f"{alogit_run_petr_path}{sheet_lookup}.F12"

    target_column_number = 9  # Excel columns are 1-indexed (1 for A, 2 for B, ...)
    lookup_column_number = target_column_number - 1

    # Read and process the text file to create a DataFrame
    with open(text_file_path, 'r') as file:
        lines = file.readlines()

    data_lines = lines[3:]
    end_index = next(i for i, line in enumerate(data_lines) if line.strip().startswith('-1'))
    relevant_lines = data_lines[:end_index]
    data = [line.split() for line in relevant_lines]
    df = pd.DataFrame(data)

    # Extract the second and fourth columns
    second_column_values = df.iloc[:, 1].values
    fourth_column_values = df.iloc[:, 3].values

    # Create a dictionary for quick lookup
    value_dict = dict(zip(second_column_values, fourth_column_values))
    used_vals = []
    # Print the vars as column if needed
    # for k in value_dict:
    #     print(k)

    sheet = workbook[sheet_name]

    # Loop through the Excel sheet to find matches and write corresponding values
    for row in sheet.iter_rows(min_row=1, max_col=lookup_column_number):
        cell_value = row[lookup_column_number - 1].value  # Get value from the lookup column
        if cell_value == None: continue
        cell_value = cell_value.split("#")[0].split("=")[0]
        if cell_value in value_dict:
            target_cell = sheet.cell(row=row[0].row, column=target_column_number)
            target_cell.value = value_dict[cell_value]
            used_vals.append(cell_value)
    print("Used vals:", used_vals)
    print("Unused vals:", [val for val in value_dict if val not in used_vals])

    # Save the workbook
    print("Saving to Excel: ", excel_file_path)
    workbook.save(excel_file_path)